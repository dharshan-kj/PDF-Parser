import os
from pdf_parser.report_generator import create_xls_report
from openpyxl import load_workbook

def test_report_generation(tmp_path):
    validation = {
        "match": False,
        "missing_titles": ["Intro"],
        "extra_titles": ["Appendix"],
        "mismatched_pages": [{"title": "Chapter 1", "expected": 5, "observed": 6}]
    }
    out = tmp_path / "report.xlsx"
    create_xls_report(validation, str(out))

    assert out.exists()
    wb = load_workbook(out)
    assert "Summary" in wb.sheetnames
    assert "Missing" in wb.sheetnames
