from openpyxl import Workbook

def create_xls_report(validation_results: dict, report_path: str):
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_missing = wb.create_sheet("Missing")
    ws_extra = wb.create_sheet("Extra")
    ws_mismatch = wb.create_sheet("Mismatched Pages")

    ws_sum.append(["All Good?", "Notes"])
    ws_sum.append([str(validation_results.get("match", False)), ""])

    ws_missing.append(["Missing Titles"])
    for t in validation_results.get("missing_titles", []):
        ws_missing.append([t])

    ws_extra.append(["Extra Titles"])
    for t in validation_results.get("extra_titles", []):
        ws_extra.append([t])

    ws_mismatch.append(["Title", "Expected Page", "Observed Page"])
    for row in validation_results.get("mismatched_pages", []):
        ws_mismatch.append([row["title"], row["expected"], row["observed"]])

    wb.save(report_path)
